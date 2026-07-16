# -*- coding: utf-8 -*-
"""Genera la base de datos del Generador de Clases de Wunda Chair."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- Datos ----------
HEADERS = [
    "ID", "Ejercicio", "Bloque sugerido", "Nivel", "Variantes",
    "Resortes (resistencia)", "Puntos de apoyo", "Grupo muscular principal",
    "Patrón de movimiento", "Objetivo", "Repeticiones sugeridas",
    "Precauciones y contraindicaciones", "Notas personales",
]

E = [
 ("Footwork: dedos (Toes)", "Calentamiento", "Básico",
  "Talones juntos, pies paralelos, una pierna",
  "Media a pesada", "Sentada en el asiento, metatarsos en el pedal, manos al costado del asiento",
  "Piernas (cuádriceps, pantorrilla) y core", "Empuje de piernas",
  "Activar piernas y conexión centro, alinear pelvis", "8 a 10",
  "Rodilla: no bloquear en extensión"),
 ("Footwork: arcos", "Calentamiento", "Básico",
  "Pies paralelos o en V",
  "Media a pesada", "Sentada, arcos del pie en el pedal, manos al costado del asiento",
  "Piernas y pie (arco plantar)", "Empuje de piernas",
  "Trabajar el arco del pie y alineación de rodilla", "8 a 10",
  "Fascitis plantar: reducir rango"),
 ("Footwork: talones", "Calentamiento", "Básico",
  "Pies paralelos, flex dorsal activa",
  "Media a pesada", "Sentada, talones en el pedal, manos al costado del asiento",
  "Isquiotibiales y glúteo", "Empuje de piernas",
  "Cargar cadena posterior sin tensar cuello", "8 a 10",
  "Ninguna relevante"),
 ("Footwork: V de Pilates", "Calentamiento", "Básico",
  "Elevando talones al final",
  "Media a pesada", "Sentada, metatarsos en el pedal en rotación externa",
  "Aductores, glúteo y piernas", "Empuje de piernas",
  "Activar línea interna de pierna y suelo pélvico", "8 a 10",
  "Cadera con pinzamiento: reducir rotación"),
 ("Pumping una pierna (sentada)", "Calentamiento", "Intermedio",
  "Pierna libre extendida o en tabletop",
  "Media", "Sentada, un pie en el pedal, pierna contraria libre, manos al costado",
  "Piernas y core (antirrotación)", "Empuje unilateral",
  "Estabilidad de pelvis con carga asimétrica", "6 a 8 por lado",
  "Inestabilidad de pelvis o dolor SI: mantener rango corto"),
 ("Frog sentada", "Calentamiento", "Básico",
  "Ritmo lento con pausa abajo",
  "Ligera a media", "Sentada al borde, metatarsos en el pedal, rodillas abiertas",
  "Aductores y suelo pélvico", "Empuje de piernas",
  "Conectar respiración, aductores y piso pélvico", "8 a 10",
  "Ninguna relevante"),
 ("Spine Stretch con pedal", "Calentamiento", "Básico",
  "Con flexión torácica mantenida, con respiración en 2 tiempos",
  "Ligera", "Sentada en el piso o banco frente a la silla, manos en el pedal, isquiones anclados",
  "Columna (movilidad en flexión) y abdomen", "Flexión de columna",
  "Movilizar columna y alargar cadena posterior", "5 a 8",
  "Osteoporosis: evitar flexión cargada; hernia discal: consultar tolerancia"),
 ("Cat Stretch arrodillada", "Calentamiento", "Básico",
  "Con pausa en redondo, con extensión neutra al subir",
  "Ligera", "Arrodillada frente a la silla, manos en el pedal, rodillas bajo cadera",
  "Columna completa y abdomen", "Flexión y extensión de columna",
  "Articular columna vértebra por vértebra", "5 a 6",
  "Muñecas sensibles: apoyar en puños o reducir carga"),
 ("Push Down de pie", "Principal", "Básico",
  "Brazos en paralelo o en V, una mano",
  "Media", "De pie frente a la silla, manos en el pedal, pies a ancho de cadera",
  "Dorsal ancho, tríceps y abdomen", "Empuje de brazos con flexión de tronco",
  "Conectar brazos al centro y estabilizar escápulas", "6 a 8",
  "Hombro inestable: reducir rango"),
 ("Pike de pie (Washerwoman)", "Principal", "Intermedio",
  "Talones elevados, una pierna atrás",
  "Media", "De pie frente a la silla, manos en el pedal, cadera alta en pica",
  "Abdomen profundo, hombros y cadena posterior", "Flexión de tronco con empuje",
  "Control del centro en posición invertida parcial", "5 a 8",
  "Hipertensión no controlada y vértigo: evitar cabeza abajo prolongado"),
 ("Pull Up", "Principal", "Avanzado",
  "Pies en V, una pierna",
  "Media a pesada", "Manos al costado del asiento, pies en el pedal, cadera alta",
  "Abdomen profundo y flexores de cadera", "Flexión de tronco (elevación de pelvis)",
  "Fuerza del centro elevando el pedal con control", "3 a 5",
  "Diástasis o prolapso sin control de presión: sustituir por Pike de pie; muñecas sensibles"),
 ("Teaser en silla", "Principal", "Avanzado",
  "Piernas dobladas, con rotación",
  "Ligera a media", "Sentada en el asiento, manos atrás en el pedal o al costado, piernas en V",
  "Abdomen completo y flexores de cadera", "Flexión de tronco con equilibrio",
  "Equilibrio y fuerza abdominal en V", "3 a 5",
  "Diástasis o prolapso: evaluar gestión de presión; lumbar agudo: sustituir"),
 ("Table Top", "Principal", "Intermedio",
  "Una pierna extendida, marcha de brazos",
  "Media", "Pies en el piso, manos en el asiento, tronco en mesa, pedal bajo manos o pies según versión",
  "Core global, glúteo y hombros", "Estabilización (plano de mesa)",
  "Sostener tronco estable contra el resorte", "20 a 30 seg",
  "Muñecas sensibles: reducir tiempo"),
 ("Mountain Climb", "Principal", "Avanzado",
  "Rango corto, con pulsos",
  "Media", "Un pie en el asiento, otro en el pedal, manos al costado del asiento, tronco inclinado",
  "Piernas, glúteo y core (equilibrio)", "Empuje unilateral en escalada",
  "Fuerza unilateral y control de pelvis en desnivel", "5 a 8 por lado",
  "Rodilla sensible: cuidar alineación; equilibrio comprometido: asistir"),
 ("Torso Twist sentada", "Principal", "Intermedio",
  "Con brazo extendido, con pausa en rotación",
  "Ligera a media", "Sentada de lado o al frente, una mano en el pedal, isquiones anclados",
  "Oblicuos y columna torácica", "Rotación de tronco",
  "Movilidad en rotación con centro activo", "4 a 6 por lado",
  "Hernia discal: rotación con rango controlado; osteoporosis: evitar rotación cargada"),
 ("Swan básico", "Principal", "Intermedio",
  "Brazos en paralelo, rango corto",
  "Ligera a media", "Abdomen sobre el asiento, manos en el pedal, piernas extendidas atrás",
  "Extensores de espalda y glúteo", "Extensión de columna",
  "Fortalecer espalda alta sin colapsar lumbar", "5 a 6",
  "Embarazo: evitar decúbito prono; espondilolistesis: evitar extensión amplia"),
 ("Swan Dive preparación", "Principal", "Avanzado",
  "Con elevación de piernas alternada",
  "Ligera a media", "Abdomen sobre el asiento, manos en el pedal, cuerpo en arco largo",
  "Cadena posterior completa", "Extensión global",
  "Extensión dinámica con control del centro", "3 a 5",
  "Lumbar agudo y espondilolistesis: sustituir por Swan básico"),
 ("Horseback", "Principal", "Avanzado",
  "Con pulsos de brazos, con pausa isométrica",
  "Ligera a media", "Sentada a horcajadas sobre el asiento, manos en el pedal, piernas abrazando la silla",
  "Aductores, suelo pélvico y abdomen", "Estabilización con empuje de brazos",
  "Conexión profunda de aductores y centro", "3 a 5",
  "Prolapso sintomático: evaluar tolerancia a la presión"),
 ("Tríceps Press sentada", "Principal", "Básico",
  "Palmas hacia atrás o hacia adentro",
  "Ligera a media", "Sentada en el piso de espaldas a la silla, manos en el pedal detrás del cuerpo",
  "Tríceps y estabilizadores de hombro", "Empuje de brazos",
  "Fuerza de brazos sin elevar hombros", "8 a 10",
  "Hombro con pinzamiento: reducir rango de extensión"),
 ("Shoulder Press de pie", "Principal", "Intermedio",
  "Una mano, ritmo con pausa",
  "Media", "De pie frente a la silla, manos en el pedal, tronco en bisagra de cadera",
  "Hombros, pecho y core", "Empuje de brazos",
  "Estabilidad escapular con tronco firme", "6 a 8",
  "Muñecas sensibles: distribuir peso en toda la palma"),
 ("Push Up sobre pedal", "Principal", "Avanzado",
  "Rodillas apoyadas, pies elevados en el asiento",
  "Media a pesada", "Manos en el pedal, cuerpo en plancha, pies en el piso o asiento",
  "Pecho, tríceps y core global", "Empuje horizontal",
  "Fuerza de empuje con plancha activa", "4 a 6",
  "Diástasis sin control: sustituir por versión con rodillas; muñecas sensibles"),
 ("Mermaid (Sirena)", "Principal", "Básico",
  "Con rotación final, con brazo libre en arco",
  "Ligera", "Sentada de lado en el asiento, mano externa en el pedal, piernas doblada en Z o pies al piso",
  "Oblicuos y cuadrado lumbar", "Flexión lateral",
  "Abrir costado y respirar hacia las costillas", "4 a 6 por lado",
  "Escoliosis: dosificar por lado según curva"),
 ("Pike lateral arrodillada", "Principal", "Avanzado",
  "Con rodilla apoyada, con pierna extendida",
  "Ligera a media", "Rodilla externa en el asiento, mano en el pedal, costado hacia el techo",
  "Oblicuos y hombro de apoyo", "Flexión lateral con empuje",
  "Fuerza lateral del tronco", "3 a 5 por lado",
  "Hombro inestable y muñeca sensible: sustituir por Mermaid"),
 ("Side Arm Press arrodillada", "Principal", "Intermedio",
  "Con pulsos, con pausa abajo",
  "Ligera", "Arrodillada de lado junto a la silla, mano en el pedal, tronco vertical",
  "Hombro (deltoides) y oblicuos", "Empuje lateral de brazo",
  "Estabilidad de tronco ante empuje unilateral", "6 a 8 por lado",
  "Rodillas sensibles: acolchar apoyo"),
 ("Going Up Front", "Principal", "Avanzado",
  "Con asistencia de manos, sin manos",
  "Media a pesada", "Un pie en el asiento, otro en el pedal, tronco vertical, manos libres o en muslo",
  "Glúteo, cuádriceps y core", "Empuje unilateral vertical (subida)",
  "Fuerza funcional de subir escalón con control", "4 a 6 por lado",
  "Rodilla: no dejar que caiga hacia adentro; equilibrio comprometido: asistir"),
 ("Step Down frontal", "Principal", "Avanzado",
  "Rango corto, con apoyo de punta",
  "Media a pesada", "De pie sobre el asiento, un pie baja al pedal, tronco vertical",
  "Glúteo y cuádriceps (excéntrico)", "Descenso unilateral controlado",
  "Control excéntrico de pierna, fuerza para bajar escaleras", "4 a 6 por lado",
  "Vértigo o equilibrio comprometido: hacer con asistencia externa"),
 ("Step Down lateral", "Principal", "Avanzado",
  "Rango corto, con brazo en oposición",
  "Media a pesada", "De pie sobre el asiento de lado, pie externo baja al pedal",
  "Glúteo medio y estabilizadores de cadera", "Descenso lateral controlado",
  "Estabilidad lateral de cadera", "4 a 6 por lado",
  "Igual que Step Down frontal"),
 ("Standing Leg Pump frontal", "Principal", "Básico",
  "Manos en cadera o al frente",
  "Media", "De pie frente a la silla, un pie en el pedal, pierna de apoyo firme",
  "Glúteo, piernas y core (equilibrio)", "Empuje unilateral de pie",
  "Equilibrio en apoyo unipodal con carga", "8 a 10 por lado",
  "Ninguna relevante; asistir si hay inestabilidad"),
 ("Standing Leg Pump lateral", "Principal", "Intermedio",
  "Con brazo externo en arco",
  "Media", "De pie de lado junto a la silla, pie interno en el pedal",
  "Aductores, glúteo medio y core", "Empuje lateral de pierna",
  "Control de pelvis en plano lateral", "6 a 8 por lado",
  "Cadera con pinzamiento: reducir rango"),
 ("Zancada con pedal (Lunge)", "Principal", "Intermedio",
  "Pulsos cortos, con inclinación de tronco",
  "Media", "Un pie en el piso adelante, el otro atrás sobre el pedal",
  "Glúteo, cuádriceps y flexores de cadera (estiramiento)", "Empuje posterior con estiramiento",
  "Fuerza de pierna delantera y apertura de cadera trasera", "6 a 8 por lado",
  "Rodilla delantera: mantener sobre el tobillo"),
 ("Achilles Stretch", "Cierre", "Básico",
  "Rodilla doblada para sóleo",
  "Ligera a media", "Metatarso en el pedal, talón buscando el piso, manos en el asiento",
  "Pantorrilla y tendón de Aquiles", "Estiramiento de cadena posterior",
  "Liberar pantorrilla y tobillo", "30 a 45 seg por lado",
  "Tendinopatía aguda de Aquiles: rango suave"),
 ("Tendon Stretch en silla", "Cierre", "Avanzado",
  "Con rodillas suaves",
  "Media", "Sentada al borde, talones en el pedal, manos al costado, flexión profunda de tronco",
  "Cadena posterior completa y abdomen", "Flexión con estiramiento activo",
  "Alargar cadena posterior con centro activo", "4 a 6",
  "Hernia discal y osteoporosis: sustituir por Achilles Stretch"),
 ("Pelvic Lift con pedal", "Cierre", "Básico",
  "Con una pierna, con pausa arriba",
  "Media", "Acostada boca arriba frente a la silla, pies en el pedal, brazos al costado",
  "Glúteo, isquiotibiales y suelo pélvico", "Extensión de cadera (puente)",
  "Integrar glúteo y respiración al cierre de la clase", "6 a 8",
  "Diástasis o posparto temprano: rango corto y exhalar al subir"),
 ("Respiración final sentada", "Cierre", "Básico",
  "Con manos en costillas, con pedal como referencia táctil",
  "Ligera o sin resorte", "Sentada en el asiento o en el piso, columna larga",
  "Diafragma y musculatura profunda", "Respiración y reorganización postural",
  "Bajar el sistema nervioso e integrar sensaciones", "5 respiraciones",
  "Ninguna"),
]

NIVELES = ["Básico", "Intermedio", "Avanzado"]
BLOQUES = ["Calentamiento", "Principal", "Cierre"]
PATRONES = sorted(set(e[7] for e in E))
GRUPOS = sorted(set(e[6] for e in E))
RESIST = ["Ligera", "Ligera a media", "Media", "Media a pesada", "Pesada", "Sin resorte"]

# ---------- Estilos ----------
FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="4A6741")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="4A6741")
SUB_FONT = Font(name=FONT, bold=True, size=11, color="4A6741")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)
ALT_FILL = PatternFill("solid", fgColor="F2F5F0")

wb = Workbook()

# ---------- Hoja 1: Ejercicios ----------
ws = wb.active
ws.title = "Ejercicios"
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center")
for i, e in enumerate(E, 2):
    row = [f"WC-{i-1:02d}", e[0], e[1], e[2], e[3], e[4], e[5], e[6], e[7], e[8], e[9], e[10], ""]
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=i, column=c, value=v)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN
        if i % 2 == 0:
            cell.fill = ALT_FILL
widths = [8, 24, 15, 12, 28, 16, 38, 26, 26, 34, 14, 42, 24]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:M{len(E)+1}"

n = len(E) + 60  # filas extra para que agregue ejercicios con validación
dv_bloque = DataValidation(type="list", formula1='"' + ",".join(BLOQUES) + '"', allow_blank=True)
dv_nivel = DataValidation(type="list", formula1='"' + ",".join(NIVELES) + '"', allow_blank=True)
dv_res = DataValidation(type="list", formula1='"' + ",".join(RESIST) + '"', allow_blank=True)
for dv, col in [(dv_bloque, "C"), (dv_nivel, "D"), (dv_res, "F")]:
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{n}")

# ---------- Hoja 2: Catálogos ----------
ws2 = wb.create_sheet("Catálogos")
cats = [("Niveles", NIVELES), ("Bloques", BLOQUES), ("Resistencias", RESIST),
        ("Patrones de movimiento", PATRONES), ("Grupos musculares", GRUPOS)]
col = 1
for name, values in cats:
    cell = ws2.cell(row=1, column=col, value=name)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    for r, v in enumerate(values, 2):
        c2 = ws2.cell(row=r, column=col, value=v)
        c2.font = BODY_FONT
        c2.border = THIN
    ws2.column_dimensions[get_column_letter(col)].width = 40
    col += 1
note = ws2.cell(row=len(max(cats, key=lambda x: len(x[1]))[1]) + 4, column=1,
    value="Puedes agregar valores nuevos aquí. Si cambias los nombres de resistencia, ajústalos también a los resortes reales de tu silla (número de resortes y posición del gancho).")
note.font = Font(name=FONT, italic=True, size=9)
ws2.merge_cells(start_row=note.row, start_column=1, end_row=note.row, end_column=4)
note.alignment = WRAP

# ---------- Hoja 3: Plantilla de clase ----------
ws3 = wb.create_sheet("Plantilla de clase")
ws3.cell(row=1, column=1, value="Estructura de clase de Wunda Chair (50 a 60 min)").font = TITLE_FONT
plan = [
    ("Bloque", "Minutos", "Ejercicios", "Criterios de selección"),
    ("Calentamiento", "8 a 10", "3 a 4",
     "Footwork siempre presente. Movilizar columna en al menos un plano. Resistencia ligera a media."),
    ("Principal 1: piernas y empuje", "10 a 12", "3 a 4",
     "Empujes de pierna sentada y de pie. Incluir al menos un ejercicio unilateral."),
    ("Principal 2: centro y flexión", "10 a 12", "3 a 4",
     "Flexión de tronco y estabilización. Progresar según nivel del grupo y contraindicaciones."),
    ("Principal 3: extensión, lateral y rotación", "10 a 12", "3 a 4",
     "Compensar la flexión con extensión (Swan), trabajo lateral (Mermaid) y una rotación."),
    ("Brazos y funcional", "6 a 8", "2 a 3",
     "Empujes de brazos y un ejercicio funcional de subida o descenso (Going Up, Step Down, Lunge)."),
    ("Cierre", "5 a 6", "2 a 3",
     "Estiramientos, Pelvic Lift y respiración final. Bajar intensidad de forma gradual."),
]
for r, row in enumerate(plan, 3):
    for c, v in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.alignment = WRAP
        cell.border = THIN
        if r == 3:
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
        else:
            cell.font = BODY_FONT
for c, w in zip(range(1, 5), [34, 10, 11, 70]):
    ws3.column_dimensions[get_column_letter(c)].width = w
rules = [
    "Reglas de balance de una buena clase:",
    "1. Toda flexión intensa se compensa con extensión en el mismo bloque o el siguiente.",
    "2. Máximo 2 ejercicios avanzados seguidos.",
    "3. Alternar posiciones (sentada, de pie, arrodillada, acostada) para transiciones fluidas.",
    "4. Revisar la columna de contraindicaciones contra el perfil de cada alumna antes de cerrar la clase.",
    "5. Los 5 patrones deben aparecer en la semana, no forzosamente en cada clase: empuje de piernas, empuje de brazos, flexión, extensión y trabajo lateral o rotación.",
]
for i, t in enumerate(rules):
    cell = ws3.cell(row=11 + i, column=1, value=t)
    cell.font = SUB_FONT if i == 0 else Font(name=FONT, size=10)
    ws3.merge_cells(start_row=11 + i, start_column=1, end_row=11 + i, end_column=4)
    cell.alignment = WRAP

# ---------- Hoja 4: Cómo usarlo ----------
ws4 = wb.create_sheet("Cómo usarlo")
ws4.column_dimensions["A"].width = 100
texts = [
    ("Generador de Clases de Wunda Chair", TITLE_FONT),
    ("", None),
    ("Qué es", SUB_FONT),
    ("Base de datos de 34 ejercicios clásicos de Wunda Chair con variantes, resortes, puntos de apoyo, nivel, patrón de movimiento, objetivo, repeticiones y contraindicaciones. Es el insumo para generar clases de 50 a 60 minutos.", None),
    ("", None),
    ("Cómo importarlo a Google Sheets", SUB_FONT),
    ("En Google Drive: Nuevo, Subir archivo, elige este xlsx y ábrelo con Google Sheets. También puedes abrir un Sheets vacío y usar Archivo, Importar.", None),
    ("", None),
    ("Cómo pedirle una clase a Claude", SUB_FONT),
    ("Comparte la hoja (o pega las filas) y pide, por ejemplo: 'Genera la clase del martes para el grupo intermedio, 55 min, foco en glúteo y espalda, Ana tiene prolapso y Marta escoliosis'. Claude filtra por nivel y contraindicaciones, arma los bloques según la Plantilla de clase y entrega la secuencia con resortes y apoyos de cada ejercicio.", None),
    ("", None),
    ("Cómo mantener la base", SUB_FONT),
    ("Agrega tus propios ejercicios en filas nuevas; las columnas Bloque, Nivel y Resortes tienen listas desplegables. Usa la columna Notas personales para tus señas de corrección o adaptaciones por alumna. Ajusta las resistencias a los resortes reales de tu silla en la hoja Catálogos.", None),
    ("", None),
    ("Importante", SUB_FONT),
    ("Las contraindicaciones son orientativas para planear la clase, no sustituyen la valoración individual de cada alumna.", None),
]
for r, (t, f) in enumerate(texts, 1):
    cell = ws4.cell(row=r, column=1, value=t)
    cell.font = f if f else BODY_FONT
    cell.alignment = WRAP

wb.save("/sessions/wonderful-cool-clarke/mnt/outputs/Generador_Clases_Wunda_Chair.xlsx")
print("OK", len(E), "ejercicios")
